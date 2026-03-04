from flask import Blueprint, request, jsonify, Response, send_file
from database import get_db
from security import require_auth
from extensions import limiter
from dotenv import load_dotenv
from sqlalchemy import func
from datetime import datetime, timedelta, timezone
from email_service import notify_status_changed, notify_new_application, notify_application_confirmation
import models
import audit
import csv
import io
import os

load_dotenv()

bp = Blueprint("candidaturas", __name__, url_prefix="/api/candidaturas")

DELIMITER     = ";"
UPLOAD_FOLDER = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "uploads"))
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
print(f"[UPLOAD] Pasta de currículos: {UPLOAD_FOLDER}")
PAGE_SIZE = 20


# ── Helpers ───────────────────────────────────────────────────

def str_to_list(value):
    if not value: return []
    return value.split(DELIMITER)

def get_list_from_data(data, field):
    value = data.get(field, "")
    if not value: return ""
    if isinstance(value, list): return DELIMITER.join(value)
    return value.replace(",", DELIMITER)

def candidatura_to_dict(c, include_history=False):
    d = {
        "id":                    c.id,
        "job": {"id": c.job.id, "position": c.job.position, "location": c.job.location},
        "fullName":              c.full_name,
        "cpf":                   c.cpf,
        "rg":                    c.rg,
        "dataNascimento":        c.data_nascimento.isoformat() if c.data_nascimento else None,
        "tipoSanguineo":         c.tipo_sanguineo,
        "nomeMae":               c.nome_mae,
        "cidadeNatal":           c.cidade_natal,
        "cidadeAtual":           c.cidade_atual,
        "phone":                 c.phone,
        "email":                 c.email,
        "linkedin":              c.linkedin,
        "education":             c.education,
        "experience":            c.experience,
        "disponibilidadeViagem": c.disponibilidade_viagem,
        "tamanhoCalca":          c.tamanho_calca,
        "tamanhoCamisa":         c.tamanho_camisa,
        "tamanhoBota":           c.tamanho_bota,
        "carteira":              str_to_list(c.carteira_motorista),
        "nrs":                   str_to_list(c.nrs),
        "escolas":               str_to_list(c.escolas),
        "motivation":            c.motivation,
        "observacoes":           c.observacoes,
        "resumeName":            c.resume_name,
        "hasResume":             bool(c.resume_name),
        "status":                c.status,
        "funnelStage":           c.funnel_stage,
        "interviewDate":         c.interview_date.isoformat() if c.interview_date else None,
        "interviewNotes":        c.interview_notes,
        "appliedAt":             c.applied_at.isoformat() if c.applied_at else None,
    }
    if include_history:
        d["history"] = [history_to_dict(h) for h in c.history]
    return d

def history_to_dict(h):
    return {
        "id":        h.id,
        "oldStatus": h.old_status,
        "newStatus": h.new_status,
        "changedBy": h.changed_by,
        "changedAt": h.changed_at.isoformat() if h.changed_at else None,
        "note":      h.note,
    }

def record_history(db, candidatura_id, old_status, new_status, changed_by, note=None):
    db.add(models.StatusHistory(
        candidatura_id=candidatura_id,
        old_status=old_status,
        new_status=new_status,
        changed_by=changed_by,
        note=note,
    ))


# ── Acompanhamento público (candidato consulta pelo CPF) ──────

@bp.get("/track")
@limiter.limit("20 per hour")
def track():
    """Candidato consulta suas candidaturas pelo CPF — sem login."""
    cpf = request.args.get("cpf", "").strip().replace(".", "").replace("-", "")
    if len(cpf) != 11 or not cpf.isdigit():
        return jsonify({"message": "CPF inválido"}), 400

    db = get_db()
    try:
        rows = (
            db.query(models.Candidatura)
            .filter(models.Candidatura.cpf.in_([cpf, f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"]))
            .order_by(models.Candidatura.applied_at.desc())
            .all()
        )
        return jsonify([{
            "id":        c.id,
            "position":  c.job.position,
            "location":  c.job.location,
            "status":    c.status,
            "appliedAt": c.applied_at.isoformat() if c.applied_at else None,
            "history": [{
                "newStatus": h.new_status,
                "changedAt": h.changed_at.isoformat() if h.changed_at else None,
                "note":      h.note,
            } for h in c.history],
        } for c in rows])
    finally:
        db.close()


# ── Submissão com rate limiting ───────────────────────────────

@bp.post("")
@limiter.limit("5 per hour")
def submit():
    if request.content_type and "multipart/form-data" in request.content_type:
        data     = request.form
        pdf_file = request.files.get("resume")
    else:
        data     = request.get_json() or {}
        pdf_file = None

    for field in ["jobId", "fullName", "cpf", "rg", "phone", "email"]:
        if not data.get(field):
            return jsonify({"message": f"Campo obrigatório: {field}"}), 400

    db = get_db()
    try:
        job = db.query(models.Job).filter_by(id=int(data["jobId"])).first()
        if not job:
            return jsonify({"message": f"Vaga não encontrada: {data['jobId']}"}), 404
        if job.status != "OPEN":
            return jsonify({"message": "Esta vaga não está mais aberta"}), 409

        duplicate = db.query(models.Candidatura).filter_by(job_id=job.id, cpf=data["cpf"]).first()
        if duplicate:
            return jsonify({"message": "CPF já cadastrado para esta vaga"}), 409

        recent_cutoff = datetime.now(timezone.utc) - timedelta(hours=24)
        if db.query(models.Candidatura).filter(
            models.Candidatura.cpf == data["cpf"],
            models.Candidatura.applied_at >= recent_cutoff,
        ).count() >= 3:
            return jsonify({"message": "Limite de candidaturas diárias atingido para este CPF."}), 429

        resume_filename = None
        if pdf_file and pdf_file.filename:
            if not pdf_file.filename.lower().endswith(".pdf"):
                return jsonify({"message": "Apenas PDFs são aceitos"}), 400
            pdf_bytes = pdf_file.read()
            if len(pdf_bytes) > 5 * 1024 * 1024:
                return jsonify({"message": "Arquivo muito grande (máx. 5 MB)"}), 400
            cpf_clean = data["cpf"].replace(".", "").replace("-", "")
            safe_name = pdf_file.filename.replace(" ", "_")
            resume_filename = f"{cpf_clean}_{job.id}_{safe_name}"
            with open(os.path.join(UPLOAD_FOLDER, resume_filename), "wb") as f:
                f.write(pdf_bytes)
            print(f"[UPLOAD] Currículo salvo: {resume_filename}")

        candidatura = models.Candidatura(
            job_id=job.id, full_name=data["fullName"], cpf=data["cpf"], rg=data["rg"],
            data_nascimento=data.get("dataNascimento") or None,
            tipo_sanguineo=data.get("tipoSanguineo") or None,
            nome_mae=data.get("nomeMae") or None,
            cidade_natal=data.get("cidadeNatal") or None,
            cidade_atual=data.get("cidadeAtual") or None,
            phone=data["phone"], email=data["email"],
            linkedin=data.get("linkedin") or None,
            education=data.get("education") or None,
            experience=data.get("experience") or None,
            disponibilidade_viagem=data.get("disponibilidadeViagem") or None,
            tamanho_calca=data.get("tamanhoCalca") or None,
            tamanho_camisa=data.get("tamanhoCamisa") or None,
            tamanho_bota=data.get("tamanhoBota") or None,
            carteira_motorista=get_list_from_data(data, "carteira"),
            nrs=get_list_from_data(data, "nrs"),
            escolas=get_list_from_data(data, "escolas"),
            motivation=data.get("motivation") or None,
            resume_name=resume_filename, status="PENDING",
        )
        db.add(candidatura)
        db.flush()

        record_history(db, candidatura.id, None, "PENDING", "sistema", "Candidatura recebida")
        db.commit()
        db.refresh(candidatura)

        audit.log("sistema", audit.NEW_APPLICATION, entity="candidatura",
                  entity_id=candidatura.id,
                  detail=f"{candidatura.full_name} → {job.position} ({job.location})")
        notify_new_application(candidatura, job)
        notify_application_confirmation(candidatura, job)  # confirmação para o candidato

        # Cria pasta no SharePoint e faz upload do currículo
        try:
            from sharepoint_service import criar_pasta_colaborador, upload_documento
            print(f"[SHAREPOINT] Criando pasta para {candidatura.full_name}...")
            result = criar_pasta_colaborador(candidatura.full_name, candidatura.cpf)
            if result.get("url"):
                print(f"[SHAREPOINT] Pasta criada: {result['url']}")
                if resume_filename:
                    pasta = result.get("pasta") or f"{candidatura.full_name} - {candidatura.cpf.replace('.','').replace('-','')}"
                    resume_path = os.path.join(UPLOAD_FOLDER, resume_filename)
                    sp_url = upload_documento(resume_path, resume_filename, pasta, sub_pasta="Curriculo")
                    if sp_url:
                        print(f"[SHAREPOINT] Curriculo enviado: {sp_url}")
                    else:
                        print(f"[SHAREPOINT] Curriculo: upload retornou None")
            else:
                print(f"[SHAREPOINT] Falha ao criar pasta: {result.get('erro')}")
        except Exception as e:
            print(f"[SHAREPOINT] Erro: {e}")

        return jsonify(candidatura_to_dict(candidatura)), 201
    finally:
        db.close()


# ── Download do currículo ─────────────────────────────────────

@bp.get("/<int:candidatura_id>/resume")
@require_auth
def download_resume(candidatura_id):
    db = get_db()
    try:
        c = db.query(models.Candidatura).filter_by(id=candidatura_id).first()
        if not c or not c.resume_name:
            return jsonify({"message": "Currículo não encontrado"}), 404
        file_path = os.path.join(UPLOAD_FOLDER, c.resume_name)
        if not os.path.exists(file_path):
            return jsonify({"message": "Arquivo não encontrado no servidor"}), 404
        audit.log(request.username, audit.DOWNLOAD_RESUME, entity="candidatura",
                  entity_id=candidatura_id, detail=f"Currículo de {c.full_name}")
        return send_file(file_path, as_attachment=True, download_name=c.resume_name)
    finally:
        db.close()


# ── Listar com filtros e paginação ────────────────────────────

@bp.get("")
@require_auth
def get_all():
    search        = request.args.get("search", "").strip()
    status_filter = request.args.get("status", "").strip().upper()
    job_filter    = request.args.get("jobId",  "").strip()
    page          = max(1, int(request.args.get("page", 1)))

    db = get_db()
    try:
        query = db.query(models.Candidatura)
        if search:
            like = f"%{search}%"
            query = query.filter(
                models.Candidatura.full_name.ilike(like) |
                models.Candidatura.cpf.ilike(like) |
                models.Candidatura.email.ilike(like) |
                models.Candidatura.phone.ilike(like)
            )
        if status_filter in ("PENDING", "APPROVED", "REJECTED"):
            query = query.filter(models.Candidatura.status == status_filter)
        if job_filter:
            query = query.filter(models.Candidatura.job_id == int(job_filter))

        total = query.count()
        rows  = query.order_by(models.Candidatura.applied_at.desc()) \
                     .offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
        return jsonify({
            "items":      [candidatura_to_dict(row) for row in rows],
            "total":      total,
            "page":       page,
            "pageSize":   PAGE_SIZE,
            "totalPages": max(1, -(-total // PAGE_SIZE)),
        })
    finally:
        db.close()


# ── Stats ─────────────────────────────────────────────────────

@bp.get("/stats")
@require_auth
def get_stats():
    db = get_db()
    try:
        return jsonify({
            "total":    db.query(models.Candidatura).count(),
            "pending":  db.query(models.Candidatura).filter_by(status="PENDING").count(),
            "approved": db.query(models.Candidatura).filter_by(status="APPROVED").count(),
            "rejected": db.query(models.Candidatura).filter_by(status="REJECTED").count(),
        })
    finally:
        db.close()


# ── Chart stats + Dashboard de tempo ─────────────────────────

@bp.get("/chart-stats")
@require_auth
def chart_stats():
    db = get_db()
    try:
        # Por vaga
        by_job = (
            db.query(models.Job.position, func.count(models.Candidatura.id).label("total"))
            .join(models.Candidatura, models.Candidatura.job_id == models.Job.id)
            .group_by(models.Job.id, models.Job.position)
            .order_by(func.count(models.Candidatura.id).desc())
            .limit(8).all()
        )

        # Por status
        status_counts = (
            db.query(models.Candidatura.status, func.count(models.Candidatura.id).label("total"))
            .group_by(models.Candidatura.status).all()
        )

        # Por mês
        all_rows = db.query(models.Candidatura.applied_at).all()
        month_map = {}
        for (dt,) in all_rows:
            if dt:
                key = dt.strftime("%Y-%m")
                month_map[key] = month_map.get(key, 0) + 1

        # ── Dashboard de tempo ────────────────────────────────
        # Candidaturas já respondidas (APPROVED ou REJECTED)
        responded = (
            db.query(models.Candidatura.id, models.Candidatura.applied_at, models.Candidatura.status)
            .filter(models.Candidatura.status.in_(["APPROVED", "REJECTED"]))
            .all()
        )

        # Busca a primeira mudança de status para calcular tempo de resposta
        response_times = []
        for c in responded:
            first_change = (
                db.query(models.StatusHistory.changed_at)
                .filter(
                    models.StatusHistory.candidatura_id == c.id,
                    models.StatusHistory.new_status.in_(["APPROVED", "REJECTED"]),
                )
                .order_by(models.StatusHistory.changed_at.asc())
                .first()
            )
            if first_change and c.applied_at:
                applied  = c.applied_at.replace(tzinfo=None) if c.applied_at.tzinfo else c.applied_at
                changed  = first_change[0].replace(tzinfo=None) if first_change[0].tzinfo else first_change[0]
                delta_h  = (changed - applied).total_seconds() / 3600
                if delta_h >= 0:
                    response_times.append(delta_h)

        avg_hours    = round(sum(response_times) / len(response_times), 1) if response_times else 0
        median_hours = sorted(response_times)[len(response_times)//2] if response_times else 0

        # Pendentes há mais tempo
        pending_old = (
            db.query(models.Candidatura.full_name, models.Candidatura.applied_at,
                     models.Job.position)
            .join(models.Job)
            .filter(models.Candidatura.status == "PENDING")
            .order_by(models.Candidatura.applied_at.asc())
            .limit(5).all()
        )

        now = datetime.now()
        oldest_pending = [{
            "name":      row.full_name,
            "position":  row.position,
            "days":      (now - row.applied_at.replace(tzinfo=None)).days if row.applied_at else 0,
        } for row in pending_old]

        # ── Comparativo entre vagas ───────────────────────────
        all_jobs = db.query(models.Job).all()
        job_comparison = []
        for job in all_jobs:
            cands = db.query(models.Candidatura).filter_by(job_id=job.id).all()
            total     = len(cands)
            approved  = sum(1 for c in cands if c.status == "APPROVED")
            rejected  = sum(1 for c in cands if c.status == "REJECTED")
            pending   = sum(1 for c in cands if c.status == "PENDING")

            # Tempo médio de resposta para esta vaga
            times = []
            for c in cands:
                if c.status in ("APPROVED", "REJECTED"):
                    fc = (
                        db.query(models.StatusHistory.changed_at)
                        .filter(
                            models.StatusHistory.candidatura_id == c.id,
                            models.StatusHistory.new_status.in_(["APPROVED", "REJECTED"]),
                        )
                        .order_by(models.StatusHistory.changed_at.asc())
                        .first()
                    )
                    if fc and c.applied_at:
                        applied = c.applied_at.replace(tzinfo=None) if c.applied_at.tzinfo else c.applied_at
                        changed = fc[0].replace(tzinfo=None) if fc[0].tzinfo else fc[0]
                        delta_h = (changed - applied).total_seconds() / 3600
                        if delta_h >= 0:
                            times.append(delta_h)

            avg_h = round(sum(times) / len(times), 1) if times else None

            if total > 0:  # só inclui vagas que têm candidatos
                job_comparison.append({
                    "id":          job.id,
                    "position":    job.position,
                    "location":    job.location,
                    "status":      job.status,
                    "total":       total,
                    "approved":    approved,
                    "rejected":    rejected,
                    "pending":     pending,
                    "approvalRate": round(approved / total * 100) if total else 0,
                    "avgHours":    avg_h,
                })

        job_comparison.sort(key=lambda x: x["total"], reverse=True)

        return jsonify({
            "byJob":         [{"label": r.position, "value": r.total} for r in by_job],
            "byMonth":       [{"label": k, "value": v} for k, v in sorted(month_map.items())[-6:]],
            "byStatus":      [{"label": r.status,   "value": r.total} for r in status_counts],
            "timeStats": {
                "avgHours":      avg_hours,
                "medianHours":   round(median_hours, 1),
                "totalResponded": len(response_times),
                "oldestPending": oldest_pending,
            },
            "jobComparison": job_comparison,
        })
    finally:
        db.close()


# ── Histórico ─────────────────────────────────────────────────

@bp.get("/<int:candidatura_id>/history")
@require_auth
def get_history(candidatura_id):
    db = get_db()
    try:
        c = db.query(models.Candidatura).filter_by(id=candidatura_id).first()
        if not c:
            return jsonify({"message": "Candidatura não encontrada"}), 404
        return jsonify([history_to_dict(h) for h in c.history])
    finally:
        db.close()


# ── Atualizar status ──────────────────────────────────────────

@bp.patch("/<int:candidatura_id>/status")
@require_auth
def update_status(candidatura_id):
    data = request.get_json()
    if not data or not data.get("status"):
        return jsonify({"message": "Status obrigatório"}), 400

    db = get_db()
    try:
        c = db.query(models.Candidatura).filter_by(id=candidatura_id).first()
        if not c:
            return jsonify({"message": "Candidatura não encontrada"}), 404

        old_status = c.status
        new_status = data["status"].upper()
        if old_status == new_status:
            return jsonify(candidatura_to_dict(c, include_history=True))

        c.status = new_status
        record_history(db, c.id, old_status, new_status,
                       request.username, data.get("note"))
        db.commit()
        db.refresh(c)

        audit.log(request.username, audit.UPDATE_STATUS, entity="candidatura",
                  entity_id=c.id,
                  detail=f"{c.full_name}: {old_status} → {new_status}")
        notify_status_changed(c, c.job)

        # Se aprovação final (APPROVED) — inicia processo de admissão
        if new_status == "APPROVED" and old_status != "APPROVED":
            try:
                from routers.processos import criar_processo_para_candidatura
                criar_processo_para_candidatura(c.id, db)
            except Exception as ex:
                print(f"[PROCESSO] Erro ao criar processo: {ex}")

        return jsonify(candidatura_to_dict(c, include_history=True))
    finally:
        db.close()


# ── Observações ───────────────────────────────────────────────

@bp.patch("/<int:candidatura_id>/observacoes")
@require_auth
def update_observacoes(candidatura_id):
    data = request.get_json()
    if data is None:
        return jsonify({"message": "Dados inválidos"}), 400
    db = get_db()
    try:
        c = db.query(models.Candidatura).filter_by(id=candidatura_id).first()
        if not c:
            return jsonify({"message": "Candidatura não encontrada"}), 404
        c.observacoes = data.get("observacoes") or None
        db.commit()
        return jsonify({"message": "Observações salvas"})
    finally:
        db.close()


# ── Exportar Excel (.xlsx) com abas por etapa ─────────────────

@bp.get("/export")
@require_auth
def export_csv():
    """
    Gera um Excel com uma aba por etapa do funil:
    Todas | Triagem | Triagem OK | Entrevista | Entrevista OK |
    Aprovação Final | Aprovados | Reprovados
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLUNAS = [
        ("ID",               lambda r: r.id),
        ("Cargo",            lambda r: r.job.position if r.job else ""),
        ("Localização",      lambda r: r.job.location if r.job else ""),
        ("Nome",             lambda r: r.full_name),
        ("CPF",              lambda r: r.cpf),
        ("RG",               lambda r: r.rg),
        ("Nascimento",       lambda r: str(r.data_nascimento) if r.data_nascimento else ""),
        ("Tipo Sanguíneo",   lambda r: r.tipo_sanguineo or ""),
        ("Nome da Mãe",      lambda r: r.nome_mae or ""),
        ("Cidade Natal",     lambda r: r.cidade_natal or ""),
        ("Cidade Atual",     lambda r: r.cidade_atual or ""),
        ("Telefone",         lambda r: r.phone),
        ("E-mail",           lambda r: r.email),
        ("LinkedIn",         lambda r: r.linkedin or ""),
        ("Formação",         lambda r: r.education or ""),
        ("Experiência",      lambda r: r.experience or ""),
        ("Disponib. Viagem", lambda r: r.disponibilidade_viagem or ""),
        ("Calça",            lambda r: r.tamanho_calca or ""),
        ("Camisa",           lambda r: r.tamanho_camisa or ""),
        ("Bota",             lambda r: r.tamanho_bota or ""),
        ("CNH",              lambda r: r.carteira_motorista or ""),
        ("NRs",              lambda r: r.nrs or ""),
        ("Escolas",          lambda r: r.escolas or ""),
        ("Motivação",        lambda r: r.motivation or ""),
        ("Observações RH",   lambda r: r.observacoes or ""),
        ("Etapa Funil",      lambda r: r.funnel_stage or ""),
        ("Status",           lambda r: r.status),
        ("Data Candidatura", lambda r: r.applied_at.strftime("%d/%m/%Y %H:%M") if r.applied_at else ""),
    ]

    # Abas: (nome_aba, filtro, cor_header_hex)
    ABAS = [
        ("📋 Todas",           None,                "#FF6A00"),
        ("🔍 Triagem",         "TRIAGEM",           "#5B8DEF"),
        ("✅ Triagem OK",      "TRIAGEM_OK",        "#2980B9"),
        ("🎤 Entrevista",      "ENTREVISTA",        "#9B59B6"),
        ("✅ Entrevista OK",   "ENTREVISTA_OK",     "#8E44AD"),
        ("🏆 Aprovação Final", "APROVACAO_FINAL",   "#F39C12"),
        ("✅ Aprovados",       "APPROVED",          "#27AE60"),
        ("❌ Reprovados",      "REJECTED",          "#E74C3C"),
    ]

    # Estilos
    def mk_header_fill(hex_color):
        c = hex_color.lstrip("#")
        return PatternFill("solid", fgColor=c)

    header_font  = Font(bold=True, color="FFFFFF", size=10)
    body_font    = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    thin_border  = Border(
        bottom=Side(style="thin", color="DDDDDD"),
    )
    alt_fill = PatternFill("solid", fgColor="1E2330")

    db = get_db()
    try:
        todas = (
            db.query(models.Candidatura)
            .order_by(models.Candidatura.applied_at.desc())
            .all()
        )

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove aba padrão

        for nome_aba, filtro, cor in ABAS:
            if filtro is None:
                rows = todas
            else:
                rows = [r for r in todas
                        if (r.funnel_stage == filtro or r.status == filtro)]

            ws = wb.create_sheet(title=nome_aba)

            # ── Cabeçalho ──────────────────────────────────────
            fill = mk_header_fill(cor)
            for col_idx, (col_name, _) in enumerate(COLUNAS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font      = header_font
                cell.fill      = fill
                cell.alignment = center_align

            ws.row_dimensions[1].height = 22

            # ── Dados ──────────────────────────────────────────
            for row_idx, cand in enumerate(rows, start=2):
                bg = PatternFill("solid", fgColor="161B27") if row_idx % 2 == 0 else alt_fill
                for col_idx, (_, getter) in enumerate(COLUNAS, start=1):
                    try:
                        val = getter(cand)
                    except Exception:
                        val = ""
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font      = body_font
                    cell.fill      = bg
                    cell.border    = thin_border
                    cell.alignment = center_align if col_idx == 1 else left_align

            # ── Largura das colunas ─────────────────────────────
            col_widths = [6, 22, 18, 28, 16, 14, 14, 10, 28, 18, 18, 16,
                          30, 28, 18, 16, 16, 8, 8, 8, 20, 20, 24, 40, 40, 18, 14, 18]
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            # ── Linha de resumo no rodapé ───────────────────────
            if rows:
                rodape_row = len(rows) + 3
                ws.cell(row=rodape_row, column=1,
                        value=f"Total: {len(rows)} candidato(s)").font = Font(
                    bold=True, color="FF6A00", size=10)

            # ── Congela primeira linha ──────────────────────────
            ws.freeze_panes = "A2"

        # ── Aba de Resumo Geral ─────────────────────────────────
        ws_res = wb.create_sheet(title="📊 Resumo", index=0)
        res_fill = mk_header_fill("#FF6A00")
        for col_idx, titulo in enumerate(["Etapa", "Total", "% do Total"], start=1):
            cell = ws_res.cell(row=1, column=col_idx, value=titulo)
            cell.font = header_font; cell.fill = res_fill
            cell.alignment = center_align
        ws_res.column_dimensions["A"].width = 28
        ws_res.column_dimensions["B"].width = 12
        ws_res.column_dimensions["C"].width = 14

        contagens = [
            ("Total Geral",     len(todas)),
            ("Triagem",         len([r for r in todas if (r.funnel_stage or r.status) == "TRIAGEM"])),
            ("Triagem OK",      len([r for r in todas if (r.funnel_stage or r.status) == "TRIAGEM_OK"])),
            ("Entrevista",      len([r for r in todas if (r.funnel_stage or r.status) == "ENTREVISTA"])),
            ("Entrevista OK",   len([r for r in todas if (r.funnel_stage or r.status) == "ENTREVISTA_OK"])),
            ("Aprovação Final", len([r for r in todas if (r.funnel_stage or r.status) == "APROVACAO_FINAL"])),
            ("Aprovados",       len([r for r in todas if r.status == "APPROVED"])),
            ("Reprovados",      len([r for r in todas if r.status == "REJECTED"])),
        ]
        total_geral = len(todas) or 1
        for i, (etapa, qtd) in enumerate(contagens, start=2):
            pct = f"{qtd/total_geral*100:.1f}%" if i > 2 else "100%"
            ws_res.cell(row=i, column=1, value=etapa).alignment = left_align
            ws_res.cell(row=i, column=2, value=qtd).alignment   = center_align
            ws_res.cell(row=i, column=3, value=pct).alignment   = center_align

        ws_res.freeze_panes = "A2"

        # ── Gera bytes e retorna ────────────────────────────────
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from datetime import date
        nome_arquivo = f"candidaturas_rezende_{date.today().strftime('%Y%m%d')}.xlsx"
        audit.log(request.username, audit.EXPORT_CSV,
                  detail=f"Exportação Excel — {len(todas)} candidaturas")

        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"},
        )
    finally:
        db.close()


# ── Funil de Seleção ──────────────────────────────────────────

FUNNEL_STAGES = [
    {"key": "TRIAGEM",          "label": "Triagem",          "next": "TRIAGEM_OK"},
    {"key": "TRIAGEM_OK",       "label": "Triagem Aprovada", "next": "ENTREVISTA"},
    {"key": "ENTREVISTA",       "label": "Entrevista",       "next": "ENTREVISTA_OK"},
    {"key": "ENTREVISTA_OK",    "label": "Entrevista Aprov.","next": "APROVACAO_FINAL"},
    {"key": "APROVACAO_FINAL",  "label": "Aprovação Final",  "next": "APPROVED"},
]


@bp.post("/<int:candidatura_id>/funnel")
@require_auth
def update_funnel(candidatura_id):
    """
    Avança ou reprova no funil de seleção.
    body: { stage, result: APPROVED|REJECTED, note, interviewDate, interviewNotes }
    """
    data   = request.get_json() or {}
    stage  = data.get("stage", "").upper()
    result = data.get("result", "").upper()   # APPROVED ou REJECTED
    note   = data.get("note", "")

    if result not in ("APPROVED", "REJECTED"):
        return jsonify({"message": "result deve ser APPROVED ou REJECTED"}), 400

    db = get_db()
    try:
        c = db.query(models.Candidatura).filter_by(id=candidatura_id).first()
        if not c:
            return jsonify({"message": "Candidatura não encontrada"}), 404

        old_stage = c.funnel_stage or "TRIAGEM"

        if result == "REJECTED":
            c.status       = "REJECTED"
            c.funnel_stage = stage
            record_history(db, c.id, old_stage, "REJECTED", request.username, note)
            db.commit()
            notify_status_changed(c, c.job)
            audit.log(request.username, "FUNNEL_REJECTED", entity="candidatura",
                      entity_id=c.id, detail=f"{c.full_name} reprovado em {stage}")
            return jsonify(candidatura_to_dict(c, include_history=True))

        # Aprovado — avança para próxima etapa
        stage_def = next((s for s in FUNNEL_STAGES if s["key"] == stage), None)
        next_stage = stage_def["next"] if stage_def else "APPROVED"

        c.funnel_stage = next_stage

        if data.get("interviewDate"):
            from datetime import datetime
            try:
                c.interview_date = datetime.fromisoformat(data["interviewDate"].replace("Z", "+00:00"))
            except Exception:
                pass
        if data.get("interviewNotes"):
            c.interview_notes = data["interviewNotes"]

        if next_stage == "APPROVED":
            c.status = "APPROVED"
            record_history(db, c.id, old_stage, "APPROVED", request.username, note)
            db.commit()
            db.refresh(c)

            # Auto-cria admissão
            try:
                import models as m
                admission = m.Admission(
                    candidatura_id = c.id,
                    full_name      = c.full_name,
                    cpf            = c.cpf,
                    cargo          = c.job.position,
                    email          = c.email,
                    phone          = c.phone,
                    status         = "IN_PROGRESS",
                    current_step   = "aso",
                    created_by     = request.username,
                )
                db.add(admission)
                db.flush()
                try:
                    from sharepoint import create_admission_folder
                    sp = create_admission_folder(c.full_name, c.cpf)
                    if sp:
                        admission.sharepoint_url = sp.get("folder_url")
                except Exception as e:
                    print(f"[SP] {e}")
                db.commit()
                audit.log(request.username, "AUTO_ADMISSION", entity="admission",
                          entity_id=admission.id,
                          detail=f"Admissão criada automaticamente para {c.full_name}")
            except Exception as e:
                print(f"[ADMISSION] Erro ao criar: {e}")

            notify_status_changed(c, c.job)
        else:
            record_history(db, c.id, old_stage, next_stage, request.username, note)
            db.commit()

        db.refresh(c)
        audit.log(request.username, "FUNNEL_ADVANCE", entity="candidatura",
                  entity_id=c.id, detail=f"{c.full_name}: {old_stage} → {next_stage}")
        return jsonify(candidatura_to_dict(c, include_history=True))
    finally:
        db.close()
